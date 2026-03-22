import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer
)
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)


def generate_pdf_report(
    group_name, currency, period_label,
    balances, settlements, expenses
):
    filename = f"report_{group_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    filepath = os.path.join("temp", filename)
    os.makedirs("temp", exist_ok=True)

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontSize=18, textColor=colors.HexColor('#1F4E79'),
        spaceAfter=12
    )
    heading_style = ParagraphStyle(
        'Heading', parent=styles['Heading2'],
        fontSize=13, textColor=colors.HexColor('#2E75B6'),
        spaceAfter=8
    )
    normal_style = styles['Normal']

    content = []

    # Title
    content.append(Paragraph("📊 SplitBazar Report", title_style))
    content.append(Paragraph(f"Group: {group_name}", normal_style))
    content.append(Paragraph(f"Period: {period_label}", normal_style))
    content.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        normal_style
    ))
    content.append(Spacer(1, 0.5*cm))

    # Summary table
    content.append(Paragraph("💰 Summary per Person", heading_style))

    summary_data = [["Name", "Paid", "Fair Share", "Balance", "Status"]]
    for user_id, data in balances.items():
        if abs(data['balance']) < 0.01:
            status = "Settled ✅"
        elif data['balance'] > 0:
            status = "Gets back 💚"
        else:
            status = "Owes ⚠️"

        summary_data.append([
            data['name'],
            f"{data['paid']:.2f} {currency}",
            f"{data['share']:.2f} {currency}",
            f"{data['balance']:+.2f} {currency}",
            status
        ])

    summary_table = Table(summary_data, colWidths=[
        4*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3*cm
    ])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#EBF3FB')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    content.append(summary_table)
    content.append(Spacer(1, 0.5*cm))

    # Settlement table
    content.append(Paragraph("💸 Final Settlement", heading_style))

    if settlements:
        settlement_data = [["From", "→", "To", "Amount"]]
        for s in settlements:
            settlement_data.append([
                s['from_name'],
                "pays",
                s['to_name'],
                f"{s['amount']:.2f} {currency}"
            ])

        settlement_table = Table(
            settlement_data,
            colWidths=[4*cm, 2*cm, 4*cm, 4*cm]
        )
        settlement_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0),
             colors.HexColor('#2E75B6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#EBF3FB')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        content.append(settlement_table)
    else:
        content.append(Paragraph(
            "✅ Everyone is settled!", normal_style
        ))

    content.append(Spacer(1, 0.5*cm))

    # Expense details
    content.append(Paragraph("🧾 Expense Details", heading_style))

    expense_data = [
        ["Date", "Paid by", "Total", "Shared",
         "Personal", "Type", "Description"]
    ]
    for exp in expenses:
        expense_data.append([
            str(exp[7]) if exp[7] else "",
            exp[8],
            f"{exp[2]:.2f}",
            f"{exp[3]:.2f}",
            f"{exp[4]:.2f}",
            exp[5] or "",
            exp[6] or ""
        ])

    expense_table = Table(expense_data, colWidths=[
        2.5*cm, 2.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 4.5*cm
    ])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0),
         colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#EBF3FB')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    content.append(expense_table)

    doc.build(content)
    return filepath


def generate_excel_report(
    group_name, currency, period_label,
    balances, settlements, expenses
):
    filename = f"report_{group_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join("temp", filename)
    os.makedirs("temp", exist_ok=True)

    wb = Workbook()

    # ── Summary Sheet ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    # Header
    header_fill = PatternFill(
        start_color="2E75B6", end_color="2E75B6",
        fill_type="solid"
    )
    header_font = Font(
        bold=True, color="FFFFFF", size=11
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws1['A1'] = "SplitBazar Report"
    ws1['A1'].font = Font(bold=True, size=16,
                          color="1F4E79")
    ws1['A2'] = f"Group: {group_name}"
    ws1['A3'] = f"Period: {period_label}"
    ws1['A4'] = f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

    ws1.append([])

    # Summary headers
    summary_headers = [
        "Name", "Paid", "Fair Share",
        "Balance", "Status"
    ]
    ws1.append(summary_headers)
    header_row = ws1.max_row
    for col in range(1, 6):
        cell = ws1.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Summary data
    for user_id, data in balances.items():
        if abs(data['balance']) < 0.01:
            status = "Settled"
        elif data['balance'] > 0:
            status = "Gets back"
        else:
            status = "Owes"

        row = [
            data['name'],
            round(data['paid'], 2),
            round(data['share'], 2),
            round(data['balance'], 2),
            status
        ]
        ws1.append(row)
        data_row = ws1.max_row
        for col in range(1, 6):
            cell = ws1.cell(row=data_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if data_row % 2 == 0:
                cell.fill = PatternFill(
                    start_color="EBF3FB",
                    end_color="EBF3FB",
                    fill_type="solid"
                )

    ws1.append([])

    # Settlement
    ws1.append(["FINAL SETTLEMENT"])
    ws1.cell(row=ws1.max_row, column=1).font = Font(
        bold=True, size=12, color="2E75B6"
    )

    if settlements:
        ws1.append(["From", "→", "To", "Amount"])
        header_row = ws1.max_row
        for col in range(1, 5):
            cell = ws1.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for s in settlements:
            ws1.append([
                s['from_name'], "pays",
                s['to_name'],
                f"{s['amount']:.2f} {currency}"
            ])
    else:
        ws1.append(["Everyone is settled!"])

    # Auto column width
    for col in ws1.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(
                    max_length, len(str(cell.value))
                )
        ws1.column_dimensions[
            col[0].column_letter
        ].width = max_length + 4

    # ── Expenses Sheet ─────────────────────────────
    ws2 = wb.create_sheet("Expenses")

    exp_headers = [
        "Date", "Paid by", "Total",
        "Shared", "Personal",
        "Split type", "Description"
    ]
    ws2.append(exp_headers)
    header_row = ws2.max_row
    for col in range(1, 8):
        cell = ws2.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for exp in expenses:
        ws2.append([
            str(exp[7]) if exp[7] else "",
            exp[8],
            float(exp[2]),
            float(exp[3]),
            float(exp[4]),
            exp[5] or "",
            exp[6] or ""
        ])
        data_row = ws2.max_row
        for col in range(1, 8):
            cell = ws2.cell(row=data_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if data_row % 2 == 0:
                cell.fill = PatternFill(
                    start_color="EBF3FB",
                    end_color="EBF3FB",
                    fill_type="solid"
                )

    for col in ws2.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(
                    max_length, len(str(cell.value))
                )
        ws2.column_dimensions[
            col[0].column_letter
        ].width = max_length + 4

    wb.save(filepath)
    return filepath